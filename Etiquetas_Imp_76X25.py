import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import os
import tempfile
import subprocess
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm
import sys

class GeneradorEtiquetasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generador de Etiquetas de Importación 76x25mm")
        self.root.geometry("700x600")
        self.root.resizable(True, True)
        
        # Información fija del importador (formato exacto como fue especificado, pero en mayúsculas)
        self.info_importador = [
            "**IMPORTADOR: **MOTORMAN DE BAJA CALIFORNIA SA DE CV",
            "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,",
            "TIJUANA, B.C. 22114 RFC: MBC210723RP9",
            "**DESCRIPCION:**",
            "**CONTENIDO:**",
            "**HECHO EN:**"
        ]
        
        # Crear el estilo para los widgets
        self.crear_estilo()
        
        # Crear la interfaz
        self.crear_interfaz()
    
    def crear_estilo(self):
        """Configura el estilo visual de la aplicación"""
        style = ttk.Style()
        
        # Configuración para botones
        style.configure('TButton', font=('Arial', 10), padding=5)
        
        # Configuración para etiquetas
        style.configure('TLabel', font=('Arial', 10))
        style.configure('Header.TLabel', font=('Arial', 14, 'bold'))
        
        # Configuración para marcos
        style.configure('TFrame', background='#f5f5f5')
        style.configure('TLabelframe', background='#f5f5f5')
        style.configure('TLabelframe.Label', font=('Arial', 11, 'bold'))
    
    def crear_interfaz(self):
        """Crea la interfaz gráfica principal"""
        # Frame principal con padding
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        titulo = ttk.Label(main_frame, text="Generador de Etiquetas de Importación 76x25mm", style='Header.TLabel')
        titulo.grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="w")
        
        # Frame para datos de etiqueta
        datos_frame = ttk.LabelFrame(main_frame, text="Datos de la Etiqueta", padding="10")
        datos_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        # Campos de entrada para datos
        ttk.Label(datos_frame, text="DESCRIPCIÓN:").grid(row=0, column=0, sticky="w", pady=5)
        self.descripcion_var = tk.StringVar(value="ABANICO PARA RADIADOR CON MOTOR")
        ttk.Entry(datos_frame, textvariable=self.descripcion_var, width=40).grid(row=0, column=1, sticky="we", pady=5, padx=5)
        
        # Entrada para CONTENIDO (solo cantidad, "PIEZA" fijo)
        ttk.Label(datos_frame, text="CONTENIDO:").grid(row=1, column=0, sticky="w", pady=5)
        
        # Frame para cantidad y unidad
        contenido_frame = ttk.Frame(datos_frame)
        contenido_frame.grid(row=1, column=1, sticky="w", pady=5, padx=5)
        
        # Spinbox para la cantidad
        self.cantidad_contenido_var = tk.IntVar(value=1)
        ttk.Spinbox(contenido_frame, from_=1, to=1000, width=5, 
                   textvariable=self.cantidad_contenido_var).pack(side=tk.LEFT)
        
        # Etiqueta fija para "PIEZA/PIEZAS"
        ttk.Label(contenido_frame, text=" PIEZA(S)").pack(side=tk.LEFT)
        
        ttk.Label(datos_frame, text="HECHO EN:").grid(row=2, column=0, sticky="w", pady=5)
        self.hecho_en_var = tk.StringVar(value="CHINA")
        ttk.Entry(datos_frame, textvariable=self.hecho_en_var, width=40).grid(row=2, column=1, sticky="we", pady=5, padx=5)
        
        ttk.Label(datos_frame, text="CANTIDAD ETIQUETAS:").grid(row=3, column=0, sticky="w", pady=5)
        self.cantidad_var = tk.IntVar(value=10)
        ttk.Spinbox(datos_frame, from_=1, to=100, textvariable=self.cantidad_var, width=10).grid(row=3, column=1, sticky="w", pady=5, padx=5)
        
        # Frame para vista previa
        preview_frame = ttk.LabelFrame(main_frame, text="Vista Previa", padding="10")
        preview_frame.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        
        # Canvas para vista previa
        self.preview_canvas = tk.Canvas(preview_frame, bg="white", width=228, height=75)  # 3x escala
        self.preview_canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Botón para actualizar vista previa
        ttk.Button(preview_frame, text="Actualizar Vista Previa", command=self.actualizar_vista_previa).pack(pady=(10, 0))
        
        # Frame para acciones
        actions_frame = ttk.LabelFrame(main_frame, text="Acciones", padding="10")
        actions_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=5, pady=10)
        
        # Botones de acción
        ttk.Button(actions_frame, text="Generar Etiquetas Excel", command=self.generar_etiquetas).grid(row=0, column=0, padx=10, pady=10)
        ttk.Button(actions_frame, text="Vista Previa de Impresión", command=self.imprimir_directamente).grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(actions_frame, text="Importar Datos desde Excel", command=self.importar_excel).grid(row=0, column=2, padx=10, pady=10)
        ttk.Button(actions_frame, text="Crear Plantilla Excel", command=self.crear_excel_ejemplo).grid(row=0, column=3, padx=10, pady=10)
        ttk.Button(actions_frame, text="Salir", command=self.root.destroy).grid(row=1, column=1, columnspan=2, padx=10, pady=10)
        
        # Configurar expansión
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Actualizar vista previa inicial
        self.actualizar_vista_previa()
    
    def actualizar_vista_previa(self):
        """Actualiza la vista previa de la etiqueta"""
        # Limpiar canvas
        self.preview_canvas.delete("all")
        
        # Dibujar borde de la etiqueta
        self.preview_canvas.create_rectangle(2, 2, 226, 73, outline="black")
        
        # Obtener datos actuales
        descripcion = self.descripcion_var.get().upper()  # Convertir a mayúsculas
        cantidad_contenido = self.cantidad_contenido_var.get()
        # Ajustar singular/plural para piezas
        if cantidad_contenido == 1:
            contenido = f"{cantidad_contenido} PIEZA"
        else:
            contenido = f"{cantidad_contenido} PIEZAS"
            
        hecho_en = self.hecho_en_var.get().upper()  # Convertir a mayúsculas
        
        # Posiciones para las líneas en la vista previa
        y_positions = [10, 20, 30, 40, 50, 60]
        
        # Tamaño de fuente para la vista previa
        font_size = 7
        
        # Dibujar cada línea
        for i, y_pos in enumerate(y_positions):
            # Para líneas con formato especial (negrita)
            if i == 0:  # Primera línea: **IMPORTADOR: **MOTORMAN...
                # Dibujar la parte "IMPORTADOR: " en negrita
                self.preview_canvas.create_text(10, y_pos, 
                                             text="IMPORTADOR: ", 
                                             font=("Arial", font_size, "bold"), 
                                             anchor="w")
                
                # Dibujar la parte "MOTORMAN..." en normal
                self.preview_canvas.create_text(80, y_pos, 
                                             text="MOTORMAN DE BAJA CALIFORNIA SA DE CV", 
                                             font=("Arial", font_size), 
                                             anchor="w")
            elif i == 1:  # Segunda línea: MARISCAL SUCRE...
                self.preview_canvas.create_text(10, y_pos, 
                                             text=self.info_importador[1], 
                                             font=("Arial", font_size), 
                                             anchor="w")
            elif i == 2:  # Tercera línea: TIJUANA, B.C. ...
                self.preview_canvas.create_text(10, y_pos, 
                                             text=self.info_importador[2], 
                                             font=("Arial", font_size), 
                                             anchor="w")
            elif i == 3:  # Cuarta línea: **DESCRIPCION:** + valor
                self.preview_canvas.create_text(10, y_pos, 
                                             text="DESCRIPCION:", 
                                             font=("Arial", font_size, "bold"), 
                                             anchor="w")
                self.preview_canvas.create_text(90, y_pos, 
                                             text=descripcion, 
                                             font=("Arial", font_size), 
                                             anchor="w")
            elif i == 4:  # Quinta línea: **CONTENIDO:** + valor
                self.preview_canvas.create_text(10, y_pos, 
                                             text="CONTENIDO:", 
                                             font=("Arial", font_size, "bold"), 
                                             anchor="w")
                self.preview_canvas.create_text(90, y_pos, 
                                             text=contenido, 
                                             font=("Arial", font_size), 
                                             anchor="w")
            elif i == 5:  # Sexta línea: **HECHO EN:** + valor
                self.preview_canvas.create_text(10, y_pos, 
                                             text="HECHO EN:", 
                                             font=("Arial", font_size, "bold"), 
                                             anchor="w")
                self.preview_canvas.create_text(90, y_pos, 
                                             text=hecho_en, 
                                             font=("Arial", font_size), 
                                             anchor="w")
    
    def crear_excel_ejemplo(self):
        """Crea un archivo Excel de ejemplo con datos de muestra"""
        try:
            # Crear datos de ejemplo
            datos = [
                {"descripcion": "ABANICO PARA RADIADOR CON MOTOR", "cantidad_contenido": 1, "hecho_en": "CHINA", "cantidad_etiquetas": 10},
                {"descripcion": "BOMBA DE AGUA", "cantidad_contenido": 2, "hecho_en": "JAPÓN", "cantidad_etiquetas": 5},
                {"descripcion": "FILTRO DE ACEITE", "cantidad_contenido": 5, "hecho_en": "MÉXICO", "cantidad_etiquetas": 20},
                {"descripcion": "SENSOR DE OXÍGENO", "cantidad_contenido": 1, "hecho_en": "CHINA", "cantidad_etiquetas": 15},
                {"descripcion": "PASTILLAS DE FRENO", "cantidad_contenido": 4, "hecho_en": "ESTADOS UNIDOS", "cantidad_etiquetas": 8},
                {"descripcion": "AMORTIGUADOR TRASERO", "cantidad_contenido": 2, "hecho_en": "TAIWÁN", "cantidad_etiquetas": 12},
            ]
            
            # Crear DataFrame y guardar como Excel
            df = pd.DataFrame(datos)
            
            # Solicitar ubicación para guardar
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Guardar plantilla Excel"
            )
            
            if file_path:
                df.to_excel(file_path, index=False)
                messagebox.showinfo("Éxito", f"Archivo de ejemplo creado: {file_path}\n\n"
                                   "La plantilla contiene las siguientes columnas:\n"
                                   "- descripcion: Descripción del producto\n"
                                   "- cantidad_contenido: Número de piezas en el paquete\n"
                                   "- hecho_en: País de fabricación\n"
                                   "- cantidad_etiquetas: Cantidad de etiquetas a generar para cada producto")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear el archivo de ejemplo: {str(e)}")
    
    def importar_excel(self):
        """Importa datos desde un archivo Excel"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Seleccionar archivo Excel"
            )
            
            if file_path:
                df = pd.read_excel(file_path)
                if df.empty:
                    messagebox.showwarning("Advertencia", "El archivo Excel está vacío")
                    return
                
                # Mostrar los datos en una nueva ventana para seleccionar
                self.mostrar_ventana_seleccion_multiple(df)
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al importar el archivo: {str(e)}")
    
    def mostrar_ventana_seleccion_multiple(self, df):
        """Muestra una ventana para seleccionar múltiples productos a importar"""
        # Crear nueva ventana
        seleccion = tk.Toplevel(self.root)
        seleccion.title("Seleccionar Productos para Etiquetas")
        seleccion.geometry("900x700")
        
        # Frame principal
        main_frame = ttk.Frame(seleccion, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Etiqueta de instrucciones
        ttk.Label(main_frame, text="Seleccione los productos para generar etiquetas:",
                 font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # Frame para la tabla
        tabla_frame = ttk.Frame(main_frame)
        tabla_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Crear Treeview para mostrar datos
        columns = list(df.columns)
        tree = ttk.Treeview(tabla_frame, columns=columns, show="headings", selectmode="extended")
        
        # Configurar encabezados
        for col in columns:
            tree.heading(col, text=col.capitalize())
            tree.column(col, width=100)
        
        # Insertar datos
        for i, row in df.iterrows():
            values = []
            for col in columns:
                if col in row:
                    values.append(row[col])
                else:
                    values.append("")
            tree.insert("", "end", iid=i, values=values)
        
        # Agregar scrollbars
        vsb = ttk.Scrollbar(tabla_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tabla_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Posicionar elementos
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        
        # Frame para botones
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        # Variables para opciones
        formato_var = tk.BooleanVar(value=True)  # True = PDF, False = Excel
        
        # Frame para opciones
        opciones_frame = ttk.LabelFrame(main_frame, text="Opciones de Generación")
        opciones_frame.pack(fill="x", pady=10)
        
        # Opciones de formato
        ttk.Radiobutton(opciones_frame, text="Generar PDF", variable=formato_var, value=True).pack(side=tk.LEFT, padx=20, pady=10)
        ttk.Radiobutton(opciones_frame, text="Generar Excel", variable=formato_var, value=False).pack(side=tk.LEFT, padx=20, pady=10)
        
        # Función para generar etiquetas para seleccionados
        def generar_para_seleccionados():
            seleccionados = tree.selection()
            if not seleccionados:
                messagebox.showwarning("Advertencia", "Debe seleccionar al menos un producto")
                return
            
            # Recopilar datos para las etiquetas
            etiquetas_datos = []
            for index in seleccionados:
                index = int(index)
                row = df.iloc[index]
                
                # Determinar cantidad de etiquetas para este producto
                cantidad_etiquetas = 1
                if "cantidad_etiquetas" in row:
                    cantidad_etiquetas = int(row["cantidad_etiquetas"])
                
                # Determinar cantidad contenido
                cantidad_contenido = 1
                if "cantidad_contenido" in row:
                    cantidad_contenido = int(row["cantidad_contenido"])
                elif "contenido" in row:
                    # Intentar extraer número del campo contenido si existe
                    contenido = str(row["contenido"])
                    try:
                        cantidad_contenido = int(''.join(filter(str.isdigit, contenido.split()[0])))
                    except:
                        cantidad_contenido = 1
                
                # Agregar tantas etiquetas como se necesiten para este producto
                for _ in range(cantidad_etiquetas):
                    etiquetas_datos.append({
                        'descripcion': row["descripcion"] if "descripcion" in row else "",
                        'cantidad_contenido': cantidad_contenido,
                        'hecho_en': row["hecho_en"] if "hecho_en" in row else ""
                    })
            
            # Guardar archivo según formato seleccionado
            formato = formato_var.get()
            
            if formato:  # PDF
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                    title="Guardar archivo PDF de etiquetas"
                )
                
                if file_path:
                    if self.generar_pdf_etiquetas(file_path, etiquetas_datos):
                        seleccion.destroy()
                        messagebox.showinfo("Éxito", f"PDF generado con {len(etiquetas_datos)} etiquetas: {file_path}")
                        
                        # Preguntar si desea abrir el PDF
                        if messagebox.askyesno("Abrir PDF", "¿Desea abrir el PDF generado?"):
                            try:
                                if os.name == 'nt':  # Windows
                                    os.startfile(file_path)
                                else:  # macOS o Linux
                                    subprocess.call(['xdg-open', file_path])
                            except:
                                messagebox.showinfo("Información", f"El PDF se ha guardado en: {file_path}")
                    
            else:  # Excel
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Guardar archivo Excel de etiquetas"
                )
                
                if file_path:
                    self.crear_etiquetas_excel(file_path, etiquetas_datos)
                    seleccion.destroy()
                    messagebox.showinfo("Éxito", f"Excel generado con {len(etiquetas_datos)} etiquetas: {file_path}")
        
        # Función para seleccionar todos
        def seleccionar_todos():
            for item in tree.get_children():
                tree.selection_add(item)
        
        # Función para deseleccionar todos
        def deseleccionar_todos():
            tree.selection_remove(tree.get_children())
        
        # Botones de acción
        ttk.Button(btn_frame, text="Generar Etiquetas", command=generar_para_seleccionados).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Seleccionar Todos", command=seleccionar_todos).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Deseleccionar Todos", command=deseleccionar_todos).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=seleccion.destroy).pack(side=tk.LEFT, padx=10)
    
    def generar_etiquetas(self):
        """Genera el archivo Excel/PDF con las etiquetas"""
        try:
            # Obtener datos actuales
            descripcion = self.descripcion_var.get()
            cantidad_contenido = self.cantidad_contenido_var.get()
            hecho_en = self.hecho_en_var.get()
            cantidad_etiquetas = self.cantidad_var.get()
            
            # Crear datos
            datos = []
            for i in range(cantidad_etiquetas):
                datos.append({
                    'descripcion': descripcion,
                    'cantidad_contenido': cantidad_contenido,
                    'hecho_en': hecho_en
                })
            
            # Preguntar el formato de salida
            formato = messagebox.askyesno("Formato de salida", 
                                         "¿Desea generar un PDF?\n\n"
                                         "Sí = Generar PDF\n"
                                         "No = Generar Excel", 
                                         icon=messagebox.QUESTION)
            
            if formato:
                # Generar PDF
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                    title="Guardar archivo PDF de etiquetas"
                )
                
                if file_path:
                    if self.generar_pdf_etiquetas(file_path, datos):
                        messagebox.showinfo("Éxito", f"Etiquetas PDF generadas: {file_path}")
                        
                        # Preguntar si desea abrir el PDF
                        if messagebox.askyesno("Abrir PDF", "¿Desea abrir el PDF generado?"):
                            try:
                                if os.name == 'nt':  # Windows
                                    os.startfile(file_path)
                                else:  # macOS o Linux
                                    subprocess.call(['xdg-open', file_path])
                            except:
                                messagebox.showinfo("Información", f"El PDF se ha guardado en: {file_path}")
            else:
                # Generar Excel
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Guardar archivo Excel de etiquetas"
                )
                
                if file_path:
                    self.crear_etiquetas_excel(file_path, datos)
                    messagebox.showinfo("Éxito", f"Etiquetas Excel generadas: {file_path}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar etiquetas: {str(e)}")
    
    def imprimir_directamente(self):
        """Genera un PDF con vista previa para impresión"""
        try:
            # Crear un archivo temporal para el PDF
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            temp_file.close()
            
            # Obtener datos actuales
            descripcion = self.descripcion_var.get()
            cantidad_contenido = self.cantidad_contenido_var.get()
            hecho_en = self.hecho_en_var.get()
            cantidad_etiquetas = self.cantidad_var.get()
            
            # Crear datos
            datos = []
            for i in range(cantidad_etiquetas):
                datos.append({
                    'descripcion': descripcion,
                    'cantidad_contenido': cantidad_contenido,
                    'hecho_en': hecho_en
                })
            
            # Mostrar mensaje antes de generar PDF
            messagebox.showinfo("Generando PDF", 
                               "Se generará un PDF con las etiquetas.\n\n"
                               "El PDF se abrirá automáticamente en su visor de PDF predeterminado,\n"
                               "donde podrá revisar las etiquetas e imprimirlas.")
            
            # Crear el archivo PDF
            if self.generar_pdf_etiquetas(temp_file.name, datos):
                # Abrir el PDF
                try:
                    if os.name == 'nt':  # Windows
                        os.startfile(temp_file.name)
                    elif os.name == 'posix' and os.path.exists('/usr/bin/open'):  # macOS
                        subprocess.call(['open', temp_file.name])
                    else:  # Linux u otros
                        subprocess.call(['xdg-open', temp_file.name])
                except Exception as e:
                    messagebox.showinfo("Información", 
                                       f"Se ha generado un PDF en: {temp_file.name}\n"
                                       "Por favor, ábralo manualmente para imprimirlo.")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al preparar la impresión: {str(e)}")
    
    def generar_pdf_etiquetas(self, archivo_salida, datos):
        """
        Genera un archivo PDF con etiquetas de importación de 76x25mm
        Formato exacto como fue especificado, todo en mayúsculas
        
        Args:
            archivo_salida: Nombre del archivo PDF a generar
            datos: Lista de diccionarios con la información de cada etiqueta
        """
        try:
            from reportlab.pdfgen import canvas
            
            # Crear un canvas para generar el PDF directamente
            c = canvas.Canvas(archivo_salida, pagesize=(76*mm, 25*mm))
            
            # Registrar fuente Arial (o usar Helvetica que viene por defecto)
            try:
                # Intenta usar Arial si está disponible
                pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
                pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
                font_name = 'Arial'
                font_name_bold = 'Arial-Bold'
            except:
                # Usa Helvetica si Arial no está disponible
                font_name = 'Helvetica'
                font_name_bold = 'Helvetica-Bold'
            
            # Para cada etiqueta
            for etiqueta in datos:
                # Determinar texto de contenido
                cantidad_contenido = etiqueta.get('cantidad_contenido', 1)
                if cantidad_contenido == 1:
                    contenido = f"{cantidad_contenido} PIEZA"
                else:
                    contenido = f"{cantidad_contenido} PIEZAS"
                
                # Convertir valores a mayúsculas
                descripcion = etiqueta['descripcion'].upper()
                hecho_en = etiqueta['hecho_en'].upper()
                
                # Ajuste de tamaño de fuente base
                font_size = 6
                
                # Posiciones para cada línea
                y_positions = [21*mm, 18*mm, 15*mm, 12*mm, 9*mm, 6*mm]
                
                # Lista de textos para cada línea
                textos = [
                    self.info_importador[0],                # **IMPORTADOR: **MOTORMAN...
                    self.info_importador[1],                # MARISCAL SUCRE...
                    self.info_importador[2],                # TIJUANA, B.C. ...
                    self.info_importador[3],                # **DESCRIPCION:**
                    self.info_importador[4],                # **CONTENIDO:**
                    self.info_importador[5]                 # **HECHO EN:**
                ]
                
                # Valores para agregar después de los encabezados
                valores = ["", "", "", descripcion, contenido, hecho_en]
                
                # Dibujar cada línea
                for i, (texto, y_pos) in enumerate(zip(textos, y_positions)):
                    # Aplicar negrita solo si el texto lo necesita
                    if "**" in texto:
                        # Descomponer el texto cuando tiene partes en negrita
                        partes = texto.split("**")
                        
                        # Calcular ancho de cada parte para alinear correctamente
                        x_pos = 5*mm  # Posición inicial desde el margen izquierdo
                        
                        for j, parte in enumerate(partes):
                            if parte:  # Si la parte no está vacía
                                if j % 2 == 1:  # Partes impares están entre ** (en negrita)
                                    c.setFont(font_name_bold, font_size)
                                else:  # Partes pares están fuera de ** (normal)
                                    c.setFont(font_name, font_size)
                                
                                # Dibujar esta parte del texto
                                c.drawString(x_pos, y_pos, parte)
                                
                                # Avanzar la posición X para la siguiente parte
                                x_pos += c.stringWidth(parte, c._fontname, font_size)
                        
                        # Si esta línea tiene un valor adicional
                        if i >= 3 and valores[i]:  # A partir de DESCRIPCION
                            c.setFont(font_name, font_size)
                            # Dibujar el valor después del encabezado
                            c.drawString(x_pos + 2*mm, y_pos, valores[i])
                    else:
                        # Texto sin formato especial - todo normal
                        c.setFont(font_name, font_size)
                        c.drawString(5*mm, y_pos, texto)
                
                # Pasar a la siguiente página
                c.showPage()
            
            # Guardar el PDF
            c.save()
            return True
            
        except Exception as e:
            import traceback
            traceback.print_exc()  # Imprimir traceback completo para debuggear
            messagebox.showerror("Error", f"Error al generar PDF: {str(e)}")
            return False
    
    def crear_etiquetas_excel(self, archivo_salida, datos):
        """
        Crea un archivo Excel con etiquetas de importación de 76x25mm
        Formato exacto como fue especificado, todo en mayúsculas
        
        Args:
            archivo_salida: Nombre del archivo Excel a generar
            datos: Lista de diccionarios con la información de cada etiqueta
        """
        # Crear un nuevo libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Etiquetas"
        
        # Definir estilos
        borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alineacion = Alignment(wrap_text=True, vertical='center', horizontal='left')
        alineacion_izquierda = Alignment(wrap_text=True, vertical='center', horizontal='left')
        negrita = Font(bold=True, size=7)  # Tamaño de fuente para las partes en negrita
        normal = Font(bold=False, size=7)  # Tamaño de fuente para texto normal
        
        # Configurar el ancho de columnas para etiquetas de 76mm (aproximadamente 25 unidades Excel)
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 25  # Aproximadamente 76mm
        
        # Configurar márgenes de página para impresión
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)
        
        # Establecer la orientación de la página horizontal para aprovechar mejor el espacio
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        
        # Crear etiquetas
        fila_actual = 1
        col_actual = 1
        
        for etiqueta in datos:
            # Calcular celda de inicio
            celda = ws.cell(row=fila_actual, column=col_actual)
            
            # Determinar texto de contenido y convertir a mayúsculas
            cantidad_contenido = etiqueta.get('cantidad_contenido', 1)
            if cantidad_contenido == 1:
                contenido = f"{cantidad_contenido} PIEZA"
            else:
                contenido = f"{cantidad_contenido} PIEZAS"
            
            descripcion = etiqueta['descripcion'].upper()
            hecho_en = etiqueta['hecho_en'].upper()
            
            # En Excel no podemos aplicar negrita a partes del texto en una celda,
            # por lo que recreamos el formato visual lo mejor posible
            linea1 = "IMPORTADOR: MOTORMAN DE BAJA CALIFORNIA SA DE CV"
            linea2 = "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,"
            linea3 = "TIJUANA, B.C. 22114 RFC: MBC210723RP9"
            linea4 = "DESCRIPCION: " + descripcion
            linea5 = "CONTENIDO: " + contenido
            linea6 = "HECHO EN: " + hecho_en
            
            # Contenido completo de la etiqueta
            contenido_etiqueta = f"{linea1}\n{linea2}\n{linea3}\n{linea4}\n{linea5}\n{linea6}"
            
            celda.value = contenido_etiqueta
            celda.font = normal  # Usar fuente normal por defecto
            celda.alignment = alineacion_izquierda  # Alineación a la izquierda
            celda.border = borde
            
            # Ajustar la altura de la fila
            ws.row_dimensions[fila_actual].height = 19
            
            # Avanzar a la siguiente posición
            col_actual += 1
            if col_actual > 3:  # 3 etiquetas por fila para aprovechar mejor el espacio
                col_actual = 1
                fila_actual += 1
        
        # Guardar el archivo
        wb.save(archivo_salida)

if __name__ == "__main__":
    root = tk.Tk()
    app = GeneradorEtiquetasApp(root)
    root.mainloop()