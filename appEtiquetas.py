import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
import io
import base64

st.set_page_config(page_title="Generador de Etiquetas", layout="wide")

# CSS personalizado
st.markdown("""
<style>
    .main {
        padding: 2rem;
    }
    .stButton > button {
        width: 100%;
        background-color: #4CAF50;
        color: white;
        padding: 0.5rem;
        border-radius: 5px;
    }
</style>
""", unsafe_allow_html=True)

# Título principal
st.title("🏷️ Generador de Etiquetas de Importación 76x25mm")
st.markdown("---")

# Información fija del importador (sin No. PARTE ya que irá en código de barras)
info_importador = [
    "**IMPORTADOR: **MOTORMAN DE BAJA CALIFORNIA SA DE CV",
    "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,",
    "TIJUANA, B.C. 22114 RFC: MBC210723RP9",
    "**DESCRIPCION:**",
    "**CONTENIDO:**",
    "**HECHO EN:**"
]

# Crear columnas para el layout
col1, col2 = st.columns([2, 1])

with col1:
    st.subheader("📝 Datos de la Etiqueta")
    
    # Campos de entrada
    descripcion = st.text_input("DESCRIPCIÓN:", value="ABANICO PARA RADIADOR CON MOTOR")
    
    # Contenido con número y unidad
    col_cont1, col_cont2 = st.columns([1, 3])
    with col_cont1:
        cantidad_contenido = st.number_input("CONTENIDO:", min_value=1, max_value=1000, value=1, step=1)
    with col_cont2:
        st.markdown("<br>", unsafe_allow_html=True)
        st.text("PIEZA(S)")
    
    hecho_en = st.text_input("HECHO EN:", value="CHINA")
    numero_parte = st.text_input("No. PARTE (SKU):", value="12345-ABC")
    cantidad_etiquetas = st.number_input("CANTIDAD DE ETIQUETAS:", min_value=1, max_value=100, value=10, step=1)

with col2:
    st.subheader("👁️ Vista Previa")
    
    # Crear vista previa con HTML
    if cantidad_contenido == 1:
        contenido_text = f"{cantidad_contenido} PIEZA"
    else:
        contenido_text = f"{cantidad_contenido} PIEZAS"
    
    preview_html = f"""
    <div style="border: 2px solid black; padding: 8px 10px; width: 228px; height: 90px; background-color: white; font-family: Arial; font-size: 9px; line-height: 1.45;">
        <div><strong>IMPORTADOR:</strong> MOTORMAN DE BAJA CALIFORNIA SA DE CV</div>
        <div>MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,</div>
        <div>TIJUANA, B.C. 22114 RFC: MBC210723RP9</div>
        <div><strong>DESCRIPCION:</strong> {descripcion.upper()}</div>
        <div><strong>CONTENIDO:</strong> {contenido_text}</div>
        <div><strong>HECHO EN:</strong> {hecho_en.upper()}</div>
    </div>
    """
    st.markdown(preview_html, unsafe_allow_html=True)

st.markdown("---")

# Funciones para generar archivos
def generar_pdf_etiquetas(datos):
    """Genera un PDF con las etiquetas (sin codigo de barras, font 8pt)."""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=(76*mm, 25*mm))

    FONT_SIZE = 8
    FONT_MIN = 5
    X_LEFT = 3*mm
    ANCHO_MAX = 70*mm
    Y_POSITIONS = [20.5*mm, 17*mm, 13.5*mm, 10*mm, 6.5*mm, 3*mm]

    def dibujar_label_valor(y, label, valor):
        fs = FONT_SIZE
        while fs > FONT_MIN:
            ancho_total = (c.stringWidth(label, "Helvetica-Bold", fs)
                           + c.stringWidth(valor, "Helvetica", fs))
            if ancho_total <= ANCHO_MAX:
                break
            fs -= 0.5
        c.setFont("Helvetica-Bold", fs)
        c.drawString(X_LEFT, y, label)
        ancho_label = c.stringWidth(label, "Helvetica-Bold", fs)
        c.setFont("Helvetica", fs)
        c.drawString(X_LEFT + ancho_label, y, valor)

    for etiqueta in datos:
        cantidad_contenido = etiqueta.get('cantidad_contenido', 1)
        contenido = f"{cantidad_contenido} PIEZA" if cantidad_contenido == 1 else f"{cantidad_contenido} PIEZAS"
        descripcion = etiqueta['descripcion'].upper()
        hecho_en = etiqueta['hecho_en'].upper()

        dibujar_label_valor(Y_POSITIONS[0], "IMPORTADOR: ", "MOTORMAN DE BAJA CALIFORNIA SA DE CV")

        c.setFont("Helvetica", FONT_SIZE)
        c.drawString(X_LEFT, Y_POSITIONS[1], "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,")
        c.drawString(X_LEFT, Y_POSITIONS[2], "TIJUANA, B.C. 22114 RFC: MBC210723RP9")

        dibujar_label_valor(Y_POSITIONS[3], "DESCRIPCION: ", descripcion)
        dibujar_label_valor(Y_POSITIONS[4], "CONTENIDO: ", contenido)
        dibujar_label_valor(Y_POSITIONS[5], "HECHO EN: ", hecho_en)

        c.showPage()

    c.save()
    buffer.seek(0)
    return buffer

def generar_excel_etiquetas(datos):
    """Genera un Excel con las etiquetas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Etiquetas"
    
    # Estilos
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    alineacion = Alignment(wrap_text=True, vertical='center', horizontal='left')
    normal = Font(bold=False, size=8)
    
    # Configurar columnas
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    # Crear etiquetas
    fila_actual = 1
    col_actual = 1
    
    for etiqueta in datos:
        celda = ws.cell(row=fila_actual, column=col_actual)
        
        # Preparar contenido
        cantidad_contenido = etiqueta.get('cantidad_contenido', 1)
        if cantidad_contenido == 1:
            contenido = f"{cantidad_contenido} PIEZA"
        else:
            contenido = f"{cantidad_contenido} PIEZAS"
        
        descripcion = etiqueta['descripcion'].upper()
        hecho_en = etiqueta['hecho_en'].upper()

        contenido_etiqueta = f"""IMPORTADOR: MOTORMAN DE BAJA CALIFORNIA SA DE CV
MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,
TIJUANA, B.C. 22114 RFC: MBC210723RP9
DESCRIPCION: {descripcion}
CONTENIDO: {contenido}
HECHO EN: {hecho_en}"""
        
        celda.value = contenido_etiqueta
        celda.font = normal
        celda.alignment = alineacion
        celda.border = borde
        
        ws.row_dimensions[fila_actual].height = 21
        
        # Avanzar posición
        col_actual += 1
        if col_actual > 3:
            col_actual = 1
            fila_actual += 1
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Sección de acciones
st.subheader("🚀 Acciones")

col_btn1, col_btn2, col_btn3 = st.columns(3)

with col_btn1:
    if st.button("📄 Generar PDF", type="primary"):
        # Preparar datos
        datos = []
        for i in range(cantidad_etiquetas):
            datos.append({
                'descripcion': descripcion,
                'cantidad_contenido': cantidad_contenido,
                'hecho_en': hecho_en,
                'numero_parte': numero_parte
            })
        
        # Generar PDF
        pdf_buffer = generar_pdf_etiquetas(datos)
        
        # Botón de descarga
        st.download_button(
            label="⬇️ Descargar PDF",
            data=pdf_buffer,
            file_name="etiquetas_importacion.pdf",
            mime="application/pdf"
        )

with col_btn2:
    if st.button("📊 Generar Excel"):
        # Preparar datos
        datos = []
        for i in range(cantidad_etiquetas):
            datos.append({
                'descripcion': descripcion,
                'cantidad_contenido': cantidad_contenido,
                'hecho_en': hecho_en,
                'numero_parte': numero_parte
            })
        
        # Generar Excel
        excel_buffer = generar_excel_etiquetas(datos)
        
        # Botón de descarga
        st.download_button(
            label="⬇️ Descargar Excel",
            data=excel_buffer,
            file_name="etiquetas_importacion.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with col_btn3:
    # Plantilla de ejemplo
    if st.button("📋 Descargar Plantilla"):
        # Crear datos de ejemplo con numero_parte
        datos_ejemplo = pd.DataFrame([
            {"descripcion": "ABANICO PARA RADIADOR CON MOTOR", "cantidad_contenido": 1, "hecho_en": "CHINA", "numero_parte": "FAN-12345", "cantidad_etiquetas": 10},
            {"descripcion": "BOMBA DE AGUA", "cantidad_contenido": 2, "hecho_en": "JAPÓN", "numero_parte": "WP-67890", "cantidad_etiquetas": 5},
            {"descripcion": "FILTRO DE ACEITE", "cantidad_contenido": 5, "hecho_en": "MÉXICO", "numero_parte": "OF-11223", "cantidad_etiquetas": 20},
        ])
        
        # Convertir a Excel
        buffer = io.BytesIO()
        datos_ejemplo.to_excel(buffer, index=False)
        buffer.seek(0)
        
        st.download_button(
            label="⬇️ Descargar Plantilla Excel",
            data=buffer,
            file_name="plantilla_etiquetas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Sección de importación desde Excel
st.markdown("---")
st.subheader("📥 Importar desde Excel")

uploaded_file = st.file_uploader("Seleccione un archivo Excel", type=['xlsx', 'xls'])

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file)
        
        st.success(f"✅ Archivo cargado: {len(df)} productos encontrados")
        
        # IMPORTANTE: Resetear el índice para evitar problemas
        df = df.reset_index(drop=True)
        
        # Mostrar datos con índice visible
        st.write("Datos cargados (con índice):")
        df_display = df.copy()
        df_display.insert(0, 'Índice', df_display.index)
        st.dataframe(df_display)
        
        # DEBUG: Mostrar información sobre el DataFrame
        with st.expander("🔍 Debug - Información del DataFrame"):
            st.write(f"Forma: {df.shape}")
            st.write(f"Columnas: {list(df.columns)}")
            st.write(f"Tipos de datos:")
            st.write(df.dtypes)
            
            # Verificar si hay valores nulos
            st.write("\nValores nulos por columna:")
            st.write(df.isnull().sum())
        
        # Selección de productos con formato mejorado
        def format_func(x):
            try:
                desc = df.loc[x, 'descripcion']
                num_parte = df.loc[x, 'numero_parte'] if 'numero_parte' in df.columns else 'Sin número'
                cant = df.loc[x, 'cantidad_etiquetas']
                return f"[{x}] {desc} | No.Parte: {num_parte} | {cant} etiquetas"
            except:
                return f"[{x}] Error al formatear"
        
        selected_indices = st.multiselect(
            "Seleccione los productos para generar etiquetas:",
            options=df.index.tolist(),
            default=df.index.tolist(),
            format_func=format_func
        )
        
        if selected_indices:
            # Mostrar lo que se va a procesar
            with st.expander("📋 Productos seleccionados para procesar"):
                for idx in selected_indices:
                    row = df.loc[idx]
                    st.write(f"**[{idx}]** {row['descripcion']} - No.Parte: {row.get('numero_parte', 'N/A')} - {row.get('cantidad_etiquetas', 1)} etiquetas")
            
            col_imp1, col_imp2, col_imp3 = st.columns(3)
            
            with col_imp1:
                if st.button("📄 Generar PDF de Seleccionados"):
                    with st.spinner("Generando PDF..."):
                        # Preparar datos
                        etiquetas_datos = []
                        
                        # Log de procesamiento
                        log_proceso = []
                        
                        # Procesar cada índice seleccionado INDIVIDUALMENTE
                        for idx in selected_indices:
                            # Obtener la fila específica
                            row = df.loc[idx]
                            
                            # Log
                            log_proceso.append(f"Procesando índice {idx}: {row['descripcion']}")
                            
                            # Obtener datos ESPECÍFICOS de ESTA fila
                            descripcion = str(row['descripcion']) if pd.notna(row['descripcion']) else ''
                            cantidad_contenido = int(row['cantidad_contenido']) if pd.notna(row.get('cantidad_contenido', 1)) else 1
                            hecho_en = str(row['hecho_en']) if pd.notna(row.get('hecho_en', '')) else ''
                            cantidad_etiquetas = int(row['cantidad_etiquetas']) if pd.notna(row.get('cantidad_etiquetas', 1)) else 1
                            
                            # IMPORTANTE: Obtener el número de parte de ESTA fila específica
                            numero_parte = ""
                            
                            # Verificar diferentes nombres de columna posibles
                            for col_name in ['numero_parte', 'no_parte', 'part_number', 'sku', 'codigo']:
                                if col_name in df.columns and pd.notna(row.get(col_name, None)):
                                    numero_parte = str(row[col_name])
                                    log_proceso.append(f"  - Número de parte encontrado en '{col_name}': {numero_parte}")
                                    break
                            
                            if not numero_parte:
                                log_proceso.append(f"  - ⚠️ No se encontró número de parte")
                            
                            # Generar las etiquetas para ESTA fila específica
                            for i in range(cantidad_etiquetas):
                                etiqueta_data = {
                                    'descripcion': descripcion,
                                    'cantidad_contenido': cantidad_contenido,
                                    'hecho_en': hecho_en,
                                    'numero_parte': numero_parte
                                }
                                etiquetas_datos.append(etiqueta_data)
                                log_proceso.append(f"  - Etiqueta {i+1}/{cantidad_etiquetas} generada con No.Parte: {numero_parte}")
                        
                        # Mostrar log de proceso
                        with st.expander("📝 Log de procesamiento"):
                            for log in log_proceso:
                                st.text(log)
                        
                        # Verificación final antes de generar PDF
                        st.info(f"Generando PDF con {len(etiquetas_datos)} etiquetas...")
                        
                        # Mostrar resumen de números de parte
                        numeros_parte_resumen = {}
                        for et in etiquetas_datos:
                            key = f"{et['descripcion'][:30]}... - {et['numero_parte']}"
                            numeros_parte_resumen[key] = numeros_parte_resumen.get(key, 0) + 1
                        
                        with st.expander("📊 Resumen de etiquetas por número de parte"):
                            for key, count in sorted(numeros_parte_resumen.items()):
                                st.write(f"{key}: {count} etiquetas")
                        
                        # Generar PDF
                        pdf_buffer = generar_pdf_etiquetas(etiquetas_datos)
                        
                        st.download_button(
                            label=f"⬇️ Descargar PDF ({len(etiquetas_datos)} etiquetas)",
                            data=pdf_buffer,
                            file_name="etiquetas_seleccionadas.pdf",
                            mime="application/pdf"
                        )
            
            with col_imp2:
                if st.button("📊 Generar Excel de Seleccionados"):
                    # Mismo proceso pero para Excel
                    etiquetas_datos = []
                    
                    for idx in selected_indices:
                        row = df.loc[idx]
                        
                        descripcion = str(row['descripcion']) if pd.notna(row['descripcion']) else ''
                        cantidad_contenido = int(row['cantidad_contenido']) if pd.notna(row.get('cantidad_contenido', 1)) else 1
                        hecho_en = str(row['hecho_en']) if pd.notna(row.get('hecho_en', '')) else ''
                        cantidad_etiquetas = int(row['cantidad_etiquetas']) if pd.notna(row.get('cantidad_etiquetas', 1)) else 1
                        
                        numero_parte = ""
                        for col_name in ['numero_parte', 'no_parte', 'part_number', 'sku', 'codigo']:
                            if col_name in df.columns and pd.notna(row.get(col_name, None)):
                                numero_parte = str(row[col_name])
                                break
                        
                        for _ in range(cantidad_etiquetas):
                            etiquetas_datos.append({
                                'descripcion': descripcion,
                                'cantidad_contenido': cantidad_contenido,
                                'hecho_en': hecho_en,
                                'numero_parte': numero_parte
                            })
                    
                    # Generar Excel
                    excel_buffer = generar_excel_etiquetas(etiquetas_datos)
                    
                    st.download_button(
                        label=f"⬇️ Descargar Excel ({len(etiquetas_datos)} etiquetas)",
                        data=excel_buffer,
                        file_name="etiquetas_seleccionadas.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            with col_imp3:
                if st.button("🐛 Ver datos procesados"):
                    st.write("Datos que se procesarían:")
                    for idx in selected_indices[:5]:  # Mostrar solo los primeros 5
                        row = df.loc[idx]
                        st.write(f"**Índice {idx}:**")
                        st.json(row.to_dict())
    
    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {str(e)}")
        st.exception(e)  # Mostrar el traceback completo

# Footer
st.markdown("---")
st.markdown("🏭 **Generador de Etiquetas de Importación** - MOTORMAN DE BAJA CALIFORNIA SA DE CV")