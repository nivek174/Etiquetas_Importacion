import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import os
import tempfile
import subprocess
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm
import sys

class GeneradorEtiquetasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generador de Etiquetas de Importación 76x51mm")
        self.root.geometry("700x600")
        self.root.resizable(True, True)
        
        # Información fija del importador (formato exacto como fue especificado, pero en mayúsculas)
        self.info_importador = [
            "**IMPORTADOR: **MOTORMAN DE BAJA CALIFORNIA SA DE CV",
            "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,",
            "TIJUANA, B.C. 22114 RFC: MBC210723RP9",
            "**DESCRIPCION:**",
            "**CONTENIDO:**",
            "**HECHO EN:**"
        ]
        
        # Crear el estilo para los widgets
        self.crear_estilo()
        
        # Crear la interfaz
        self.crear_interfaz()
    
    def crear_estilo(self):
        """Configura el estilo visual de la aplicación"""
        style = ttk.Style()
        
        # Configuración para botones
        style.configure('TButton', font=('Arial', 10), padding=5)
        
        # Configuración para etiquetas
        style.configure('TLabel', font=('Arial', 10))
        style.configure('Header.TLabel', font=('Arial', 14, 'bold'))
        
        # Configuración para marcos
        style.configure('TFrame', background='#f5f5f5')
        style.configure('TLabelframe', background='#f5f5f5')
        style.configure('TLabelframe.Label', font=('Arial', 11, 'bold'))
    
    def crear_interfaz(self):
        """Crea la interfaz gráfica principal"""
        # Frame principal con padding
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        titulo = ttk.Label(main_frame, text="Generador de Etiquetas de Importación 76x51mm", style='Header.TLabel')
        titulo.grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="w")
        
        # Frame para datos de etiqueta
        datos_frame = ttk.LabelFrame(main_frame, text="Datos de la Etiqueta", padding="10")
        datos_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        # Campos de entrada para datos
        ttk.Label(datos_frame, text="DESCRIPCIÓN:").grid(row=0, column=0, sticky="w", pady=5)
        self.descripcion_var = tk.StringVar(value="ABANICO PARA RADIADOR CON MOTOR")
        ttk.Entry(datos_frame, textvariable=self.descripcion_var, width=40).grid(row=0, column=1, sticky="we", pady=5, padx=5)
        
        # Entrada para CONTENIDO (solo cantidad, "PIEZA" fijo)
        ttk.Label(datos_frame, text="CONTENIDO:").grid(row=1, column=0, sticky="w", pady=5)
        
        # Frame para cantidad y unidad
        contenido_frame = ttk.Frame(datos_frame)
        contenido_frame.grid(row=1, column=1, sticky="w", pady=5, padx=5)
        
        # Spinbox para la cantidad
        self.cantidad_contenido_var = tk.IntVar(value=1)
        ttk.Spinbox(contenido_frame, from_=1, to=1000, width=5, 
                   textvariable=self.cantidad_contenido_var).pack(side=tk.LEFT)
        
        # Etiqueta fija para "PIEZA/PIEZAS"
        ttk.Label(contenido_frame, text=" PIEZA(S)").pack(side=tk.LEFT)
        
        ttk.Label(datos_frame, text="HECHO EN:").grid(row=2, column=0, sticky="w", pady=5)
        self.hecho_en_var = tk.StringVar(value="CHINA")
        ttk.Entry(datos_frame, textvariable=self.hecho_en_var, width=40).grid(row=2, column=1, sticky="we", pady=5, padx=5)
        
        ttk.Label(datos_frame, text="CANTIDAD ETIQUETAS:").grid(row=3, column=0, sticky="w", pady=5)
        self.cantidad_var = tk.IntVar(value=10)
        ttk.Spinbox(datos_frame, from_=1, to=100, textvariable=self.cantidad_var, width=10).grid(row=3, column=1, sticky="w", pady=5, padx=5)
        
        # Frame para vista previa
        preview_frame = ttk.LabelFrame(main_frame, text="Vista Previa", padding="10")
        preview_frame.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        
        # Canvas para vista previa (ajustado para 76x51mm - aproximadamente el doble de alto)
        self.preview_canvas = tk.Canvas(preview_frame, bg="white", width=228, height=153)  # 3x escala
        self.preview_canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Botón para actualizar vista previa
        ttk.Button(preview_frame, text="Actualizar Vista Previa", command=self.actualizar_vista_previa).pack(pady=(10, 0))
        
        # Frame para acciones
        actions_frame = ttk.LabelFrame(main_frame, text="Acciones", padding="10")
        actions_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=5, pady=10)
        
        # Botones de acción
        ttk.Button(actions_frame, text="Generar Etiquetas Excel", command=self.generar_etiquetas).grid(row=0, column=0, padx=10, pady=10)
        ttk.Button(actions_frame, text="Vista Previa de Impresión", command=self.imprimir_directamente).grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(actions_frame, text="Importar Datos desde Excel", command=self.importar_excel).grid(row=0, column=2, padx=10, pady=10)
        ttk.Button(actions_frame, text="Crear Plantilla Excel", command=self.crear_excel_ejemplo).grid(row=0, column=3, padx=10, pady=10)
        ttk.Button(actions_frame, text="Salir", command=self.root.destroy).grid(row=1, column=1, columnspan=2, padx=10, pady=10)
        
        # Configurar expansión
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Actualizar vista previa inicial
        self.actualizar_vista_previa()
    
    def actualizar_vista_previa(self):
        """Actualiza la vista previa de la etiqueta"""
        # Limpiar canvas
        self.preview_canvas.delete("all")
        
        # Dibujar borde de la etiqueta
        self.preview_canvas.create_rectangle(2, 2, 226, 151, outline="black")
        
        # Obtener datos actuales
        descripcion = self.descripcion_var.get().upper()  # Convertir a mayúsculas
        cantidad_contenido = self.cantidad_contenido_var.get()
        # Ajustar singular/plural para piezas
        if cantidad_contenido == 1:
            contenido = f"{cantidad_contenido} PIEZA"
        else:
            contenido = f"{cantidad_contenido} PIEZAS"
            
        hecho_en = self.hecho_en_var.get().upper()  # Convertir a mayúsculas
        
        # Posiciones para las líneas en la vista previa (ahora con más espacio vertical)
        y_positions = [20, 40, 60, 85, 110, 135]
        
        # Tamaño de fuente base para la vista previa
        font_size_base = 10
        
        # Ancho máximo disponible para el texto (en pixeles)
        max_width = 210  # Un poco menos que el ancho total de la etiqueta (226px)
        
        # Posición X inicial para todos los textos
        x_inicial = 10
        
        # Posición X para los valores (alineados verticalmente)
        x_valores = 120  # Esto hace que todos los valores estén alineados a la misma distancia
        
        # Función para medir el ancho real de un texto con una fuente específica
        def medir_ancho_texto(texto, tamano, negrita=False):
            font_style = "bold" if negrita else ""
            font = (f"Arial {tamano} {font_style}") if negrita else (f"Arial {tamano}")
            width = self.preview_canvas.winfo_toplevel().winfo_fpixels('1i') * len(texto) * tamano / 72
            return width
        
        # Dibujar línea 1: IMPORTADOR: MOTORMAN DE BAJA CALIFORNIA SA DE CV
        importador_bold = "IMPORTADOR:"
        importador_normal = "MOTORMAN DE BAJA CALIFORNIA SA DE CV"
        
        # Calcular tamaño para título
        tamano_titulo = font_size_base
        self.preview_canvas.create_text(x_inicial, y_positions[0], 
                                     text=importador_bold, 
                                     font=("Arial", tamano_titulo, "bold"),
                                     anchor="w")
        
        # Calcular tamaño para valor (considerando espacio disponible)
        tamano_valor = font_size_base
        ancho_valor = medir_ancho_texto(importador_normal, tamano_valor)
        espacio_disponible = max_width - x_valores + x_inicial
        
        while ancho_valor > espacio_disponible and tamano_valor > 6:
            tamano_valor -= 1
            ancho_valor = medir_ancho_texto(importador_normal, tamano_valor)
        
        # Dibujar valor alineado con posición fija
        self.preview_canvas.create_text(x_valores, y_positions[0],
                                     text=importador_normal,
                                     font=("Arial", tamano_valor),
                                     anchor="w")
        
        # Línea 2: MARISCAL SUCRE...
        direccion1 = "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,"
        tamano_dir = tamano_valor  # Mismo tamaño que el texto anterior
        ancho_dir1 = medir_ancho_texto(direccion1, tamano_dir)
        
        while ancho_dir1 > max_width - x_inicial and tamano_dir > 6:
            tamano_dir -= 1
            ancho_dir1 = medir_ancho_texto(direccion1, tamano_dir)
        
        self.preview_canvas.create_text(x_inicial, y_positions[1],
                                     text=direccion1,
                                     font=("Arial", tamano_dir),
                                     anchor="w")
        
        # Línea 3: TIJUANA, B.C. ...
        direccion2 = "TIJUANA, B.C. 22114 RFC: MBC210723RP9"
        self.preview_canvas.create_text(x_inicial, y_positions[2],
                                     text=direccion2,
                                     font=("Arial", tamano_dir),
                                     anchor="w")
        
        # Línea 4: DESCRIPCION: valor
        desc_bold = "DESCRIPCION:"
        tamano_desc_titulo = tamano_titulo
        
        self.preview_canvas.create_text(x_inicial, y_positions[3],
                                     text=desc_bold,
                                     font=("Arial", tamano_desc_titulo, "bold"),
                                     anchor="w")
        
        # Calcular tamaño para valor
        tamano_desc_valor = tamano_valor
        ancho_desc_valor = medir_ancho_texto(descripcion, tamano_desc_valor)
        
        while ancho_desc_valor > max_width - x_valores + x_inicial and tamano_desc_valor > 6:
            tamano_desc_valor -= 1
            ancho_desc_valor = medir_ancho_texto(descripcion, tamano_desc_valor)
        
        self.preview_canvas.create_text(x_valores, y_positions[3],
                                     text=descripcion,
                                     font=("Arial", tamano_desc_valor),
                                     anchor="w")
        
        # Línea 5: CONTENIDO: valor
        cont_bold = "CONTENIDO:"
        self.preview_canvas.create_text(x_inicial, y_positions[4],
                                     text=cont_bold,
                                     font=("Arial", tamano_titulo, "bold"),
                                     anchor="w")
        
        self.preview_canvas.create_text(x_valores, y_positions[4],
                                     text=contenido,
                                     font=("Arial", tamano_valor),
                                     anchor="w")
        
        # Línea 6: HECHO EN: valor
        hecho_bold = "HECHO EN:"
        self.preview_canvas.create_text(x_inicial, y_positions[5],
                                     text=hecho_bold,
                                     font=("Arial", tamano_titulo, "bold"),
                                     anchor="w")
        
        # Calcular tamaño para valor
        tamano_hecho = tamano_valor
        ancho_hecho = medir_ancho_texto(hecho_en, tamano_hecho)
        
        while ancho_hecho > max_width - x_valores + x_inicial and tamano_hecho > 6:
            tamano_hecho -= 1
            ancho_hecho = medir_ancho_texto(hecho_en, tamano_hecho)
        
        self.preview_canvas.create_text(x_valores, y_positions[5],
                                     text=hecho_en,
                                     font=("Arial", tamano_hecho),
                                     anchor="w")
    
    def crear_excel_ejemplo(self):
        """Crea un archivo Excel de ejemplo con datos de muestra"""
        try:
            # Crear datos de ejemplo
            datos = [
                {"descripcion": "ABANICO PARA RADIADOR CON MOTOR", "cantidad_contenido": 1, "hecho_en": "CHINA", "cantidad_etiquetas": 10},
                {"descripcion": "BOMBA DE AGUA", "cantidad_contenido": 2, "hecho_en": "JAPÓN", "cantidad_etiquetas": 5},
                {"descripcion": "FILTRO DE ACEITE", "cantidad_contenido": 5, "hecho_en": "MÉXICO", "cantidad_etiquetas": 20},
                {"descripcion": "SENSOR DE OXÍGENO", "cantidad_contenido": 1, "hecho_en": "CHINA", "cantidad_etiquetas": 15},
                {"descripcion": "PASTILLAS DE FRENO", "cantidad_contenido": 4, "hecho_en": "ESTADOS UNIDOS", "cantidad_etiquetas": 8},
                {"descripcion": "AMORTIGUADOR TRASERO", "cantidad_contenido": 2, "hecho_en": "TAIWÁN", "cantidad_etiquetas": 12},
            ]
            
            # Crear DataFrame y guardar como Excel
            df = pd.DataFrame(datos)
            
            # Solicitar ubicación para guardar
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Guardar plantilla Excel"
            )
            
            if file_path:
                df.to_excel(file_path, index=False)
                messagebox.showinfo("Éxito", f"Archivo de ejemplo creado: {file_path}\n\n"
                                   "La plantilla contiene las siguientes columnas:\n"
                                   "- descripcion: Descripción del producto\n"
                                   "- cantidad_contenido: Número de piezas en el paquete\n"
                                   "- hecho_en: País de fabricación\n"
                                   "- cantidad_etiquetas: Cantidad de etiquetas a generar para cada producto")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear el archivo de ejemplo: {str(e)}")
    
    def importar_excel(self):
        """Importa datos desde un archivo Excel"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Seleccionar archivo Excel"
            )
            
            if file_path:
                df = pd.read_excel(file_path)
                if df.empty:
                    messagebox.showwarning("Advertencia", "El archivo Excel está vacío")
                    return
                
                # Mostrar los datos en una nueva ventana para seleccionar
                self.mostrar_ventana_seleccion_multiple(df)
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al importar el archivo: {str(e)}")
    
    def mostrar_ventana_seleccion_multiple(self, df):
        """Muestra una ventana para seleccionar múltiples productos a importar"""
        # Crear nueva ventana
        seleccion = tk.Toplevel(self.root)
        seleccion.title("Seleccionar Productos para Etiquetas")
        seleccion.geometry("900x700")
        
        # Frame principal
        main_frame = ttk.Frame(seleccion, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Etiqueta de instrucciones
        ttk.Label(main_frame, text="Seleccione los productos para generar etiquetas:",
                 font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # Frame para la tabla
        tabla_frame = ttk.Frame(main_frame)
        tabla_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Crear Treeview para mostrar datos
        columns = list(df.columns)
        tree = ttk.Treeview(tabla_frame, columns=columns, show="headings", selectmode="extended")
        
        # Configurar encabezados
        for col in columns:
            tree.heading(col, text=col.capitalize())
            tree.column(col, width=100)
        
        # Insertar datos
        for i, row in df.iterrows():
            values = []
            for col in columns:
                if col in row:
                    values.append(row[col])
                else:
                    values.append("")
            tree.insert("", "end", iid=i, values=values)
        
        # Agregar scrollbars
        vsb = ttk.Scrollbar(tabla_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tabla_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Posicionar elementos
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        
        # Frame para botones
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        # Variables para opciones
        formato_var = tk.BooleanVar(value=True)  # True = PDF, False = Excel
        
        # Frame para opciones
        opciones_frame = ttk.LabelFrame(main_frame, text="Opciones de Generación")
        opciones_frame.pack(fill="x", pady=10)
        
        # Opciones de formato
        ttk.Radiobutton(opciones_frame, text="Generar PDF", variable=formato_var, value=True).pack(side=tk.LEFT, padx=20, pady=10)
        ttk.Radiobutton(opciones_frame, text="Generar Excel", variable=formato_var, value=False).pack(side=tk.LEFT, padx=20, pady=10)
        
        # Función para generar etiquetas para seleccionados
        def generar_para_seleccionados():
            seleccionados = tree.selection()
            if not seleccionados:
                messagebox.showwarning("Advertencia", "Debe seleccionar al menos un producto")
                return
            
            # Recopilar datos para las etiquetas
            etiquetas_datos = []
            for index in seleccionados:
                index = int(index)
                row = df.iloc[index]
                
                # Determinar cantidad de etiquetas para este producto
                cantidad_etiquetas = 1
                if "cantidad_etiquetas" in row:
                    cantidad_etiquetas = int(row["cantidad_etiquetas"])
                
                # Determinar cantidad contenido
                cantidad_contenido = 1
                if "cantidad_contenido" in row:
                    cantidad_contenido = int(row["cantidad_contenido"])
                elif "contenido" in row:
                    # Intentar extraer número del campo contenido si existe
                    contenido = str(row["contenido"])
                    try:
                        cantidad_contenido = int(''.join(filter(str.isdigit, contenido.split()[0])))
                    except:
                        cantidad_contenido = 1
                
                # Agregar tantas etiquetas como se necesiten para este producto
                for _ in range(cantidad_etiquetas):
                    etiquetas_datos.append({
                        'descripcion': row["descripcion"] if "descripcion" in row else "",
                        'cantidad_contenido': cantidad_contenido,
                        'hecho_en': row["hecho_en"] if "hecho_en" in row else ""
                    })
            
            # Guardar archivo según formato seleccionado
            formato = formato_var.get()
            
            if formato:  # PDF
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                    title="Guardar archivo PDF de etiquetas"
                )
                
                if file_path:
                    if self.generar_pdf_etiquetas(file_path, etiquetas_datos):
                        seleccion.destroy()
                        messagebox.showinfo("Éxito", f"PDF generado con {len(etiquetas_datos)} etiquetas: {file_path}")
                        
                        # Preguntar si desea abrir el PDF
                        if messagebox.askyesno("Abrir PDF", "¿Desea abrir el PDF generado?"):
                            try:
                                if os.name == 'nt':  # Windows
                                    os.startfile(file_path)
                                else:  # macOS o Linux
                                    subprocess.call(['xdg-open', file_path])
                            except:
                                messagebox.showinfo("Información", f"El PDF se ha guardado en: {file_path}")
                    
            else:  # Excel
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Guardar archivo Excel de etiquetas"
                )
                
                if file_path:
                    self.crear_etiquetas_excel(file_path, etiquetas_datos)
                    seleccion.destroy()
                    messagebox.showinfo("Éxito", f"Excel generado con {len(etiquetas_datos)} etiquetas: {file_path}")
        
        # Función para seleccionar todos
        def seleccionar_todos():
            for item in tree.get_children():
                tree.selection_add(item)
        
        # Función para deseleccionar todos
        def deseleccionar_todos():
            tree.selection_remove(tree.get_children())
        
        # Botones de acción
        ttk.Button(btn_frame, text="Generar Etiquetas", command=generar_para_seleccionados).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Seleccionar Todos", command=seleccionar_todos).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Deseleccionar Todos", command=deseleccionar_todos).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=seleccion.destroy).pack(side=tk.LEFT, padx=10)
    
    def generar_etiquetas(self):
        """Genera el archivo Excel/PDF con las etiquetas"""
        try:
            # Obtener datos actuales
            descripcion = self.descripcion_var.get()
            cantidad_contenido = self.cantidad_contenido_var.get()
            hecho_en = self.hecho_en_var.get()
            cantidad_etiquetas = self.cantidad_var.get()
            
            # Crear datos
            datos = []
            for i in range(cantidad_etiquetas):
                datos.append({
                    'descripcion': descripcion,
                    'cantidad_contenido': cantidad_contenido,
                    'hecho_en': hecho_en
                })
            
            # Preguntar el formato de salida
            formato = messagebox.askyesno("Formato de salida", 
                                         "¿Desea generar un PDF?\n\n"
                                         "Sí = Generar PDF\n"
                                         "No = Generar Excel", 
                                         icon=messagebox.QUESTION)
            
            if formato:
                # Generar PDF
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                    title="Guardar archivo PDF de etiquetas"
                )
                
                if file_path:
                    if self.generar_pdf_etiquetas(file_path, datos):
                        messagebox.showinfo("Éxito", f"Etiquetas PDF generadas: {file_path}")
                        
                        # Preguntar si desea abrir el PDF
                        if messagebox.askyesno("Abrir PDF", "¿Desea abrir el PDF generado?"):
                            try:
                                if os.name == 'nt':  # Windows
                                    os.startfile(file_path)
                                else:  # macOS o Linux
                                    subprocess.call(['xdg-open', file_path])
                            except:
                                messagebox.showinfo("Información", f"El PDF se ha guardado en: {file_path}")
            else:
                # Generar Excel
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Guardar archivo Excel de etiquetas"
                )
                
                if file_path:
                    self.crear_etiquetas_excel(file_path, datos)
                    messagebox.showinfo("Éxito", f"Etiquetas Excel generadas: {file_path}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar etiquetas: {str(e)}")
    
    def imprimir_directamente(self):
        """Genera un PDF con vista previa para impresión"""
        try:
            # Crear un archivo temporal para el PDF
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            temp_file.close()
            
            # Obtener datos actuales
            descripcion = self.descripcion_var.get()
            cantidad_contenido = self.cantidad_contenido_var.get()
            hecho_en = self.hecho_en_var.get()
            cantidad_etiquetas = self.cantidad_var.get()
            
            # Crear datos
            datos = []
            for i in range(cantidad_etiquetas):
                datos.append({
                    'descripcion': descripcion,
                    'cantidad_contenido': cantidad_contenido,
                    'hecho_en': hecho_en
                })
            
            # Mostrar mensaje antes de generar PDF
            messagebox.showinfo("Generando PDF", 
                               "Se generará un PDF con las etiquetas.\n\n"
                               "El PDF se abrirá automáticamente en su visor de PDF predeterminado,\n"
                               "donde podrá revisar las etiquetas e imprimirlas.")
            
            # Crear el archivo PDF
            if self.generar_pdf_etiquetas(temp_file.name, datos):
                # Abrir el PDF
                try:
                    if os.name == 'nt':  # Windows
                        os.startfile(temp_file.name)
                    elif os.name == 'posix' and os.path.exists('/usr/bin/open'):  # macOS
                        subprocess.call(['open', temp_file.name])
                    else:  # Linux u otros
                        subprocess.call(['xdg-open', temp_file.name])
                except Exception as e:
                    messagebox.showinfo("Información", 
                                       f"Se ha generado un PDF en: {temp_file.name}\n"
                                       "Por favor, ábralo manualmente para imprimirlo.")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al preparar la impresión: {str(e)}")
    
    def generar_pdf_etiquetas(self, archivo_salida, datos):
        """
        Genera un archivo PDF con etiquetas de importación de 76x51mm
        Formato exacto como fue especificado, todo en mayúsculas
        
        Args:
            archivo_salida: Nombre del archivo PDF a generar
            datos: Lista de diccionarios con la información de cada etiqueta
        """
        try:
            from reportlab.pdfgen import canvas
            
            # Crear un canvas para generar el PDF directamente
            c = canvas.Canvas(archivo_salida, pagesize=(76*mm, 51*mm))
            
            # Registrar fuente Arial (o usar Helvetica que viene por defecto)
            try:
                # Intenta usar Arial si está disponible
                pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
                pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
                font_name = 'Arial'
                font_name_bold = 'Arial-Bold'
            except:
                # Usa Helvetica si Arial no está disponible
                font_name = 'Helvetica'
                font_name_bold = 'Helvetica-Bold'
            
            # Para cada etiqueta
            for etiqueta in datos:
                # Determinar texto de contenido
                cantidad_contenido = etiqueta.get('cantidad_contenido', 1)
                if cantidad_contenido == 1:
                    contenido = f"{cantidad_contenido} PIEZA"
                else:
                    contenido = f"{cantidad_contenido} PIEZAS"
                
                # Convertir valores a mayúsculas
                descripcion = etiqueta['descripcion'].upper()
                hecho_en = etiqueta['hecho_en'].upper()
                
                # Posiciones para cada línea (ajustadas para 51mm de altura)
                y_positions = [45*mm, 39*mm, 33*mm, 27*mm, 21*mm, 15*mm]
                
                # Datos fijos
                importador_bold = "IMPORTADOR:"
                importador_normal = "MOTORMAN DE BAJA CALIFORNIA SA DE CV"
                direccion1 = "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,"
                direccion2 = "TIJUANA, B.C. 22114 RFC: MBC210723RP9"
                descripcion_bold = "DESCRIPCION:"
                contenido_bold = "CONTENIDO:"
                hecho_en_bold = "HECHO EN:"
                
                # Ajuste de tamaño de fuente base para etiquetas 76x51mm
                font_size_base = 9  # Tamaño de base
                
                # Ancho máximo disponible en mm (margen de 5mm a cada lado)
                max_ancho = 66  # 76mm - 10mm de márgenes
                
                # Posición X inicial para todos los textos
                x_inicial = 5*mm
                
                # Posición X para los valores (alineados verticalmente)
                x_valores = 31*mm  # Posición fija para todos los valores
                
                # Dibujar línea 1: IMPORTADOR: MOTORMAN DE BAJA CALIFORNIA SA DE CV
                c.setFont(font_name_bold, font_size_base)
                c.drawString(x_inicial, y_positions[0], importador_bold)
                
                # Dibujar el valor con tamaño ajustado si es necesario
                c.setFont(font_name, font_size_base)
                ancho_importador_normal = c.stringWidth(importador_normal, font_name, font_size_base)
                espacio_disponible = max_ancho*mm - x_valores + x_inicial
                
                tamano_valor = font_size_base
                while ancho_importador_normal > espacio_disponible and tamano_valor > 5.5:
                    tamano_valor -= 0.5
                    c.setFont(font_name, tamano_valor)
                    ancho_importador_normal = c.stringWidth(importador_normal, font_name, tamano_valor)
                
                c.drawString(x_valores, y_positions[0], importador_normal)
                
                # Línea 2: Dirección línea 1
                c.setFont(font_name, font_size_base)
                ancho_direccion1 = c.stringWidth(direccion1, font_name, font_size_base)
                tamano_dir = font_size_base
                
                # Reducir tamaño si es necesario
                while ancho_direccion1 > max_ancho*mm and tamano_dir > 5.5:
                    tamano_dir -= 0.5
                    c.setFont(font_name, tamano_dir)
                    ancho_direccion1 = c.stringWidth(direccion1, font_name, tamano_dir)
                
                c.drawString(x_inicial, y_positions[1], direccion1)
                
                # Línea 3: Dirección línea 2
                c.setFont(font_name, tamano_dir)  # Mismo tamaño que la línea anterior
                c.drawString(x_inicial, y_positions[2], direccion2)
                
                # Línea 4: DESCRIPCION: + valor
                c.setFont(font_name_bold, font_size_base)
                c.drawString(x_inicial, y_positions[3], descripcion_bold)
                
                # Calcular tamaño adecuado para la descripción
                c.setFont(font_name, font_size_base)
                ancho_descripcion = c.stringWidth(descripcion, font_name, font_size_base)
                tamano_desc = font_size_base
                
                while ancho_descripcion > espacio_disponible and tamano_desc > 5.5:
                    tamano_desc -= 0.5
                    c.setFont(font_name, tamano_desc)
                    ancho_descripcion = c.stringWidth(descripcion, font_name, tamano_desc)
                
                c.drawString(x_valores, y_positions[3], descripcion)
                
                # Línea 5: CONTENIDO: + valor
                c.setFont(font_name_bold, font_size_base)
                c.drawString(x_inicial, y_positions[4], contenido_bold)
                
                # El contenido suele ser corto, no necesita ajuste
                c.setFont(font_name, font_size_base)
                c.drawString(x_valores, y_positions[4], contenido)
                
                # Línea 6: HECHO EN: + valor
                c.setFont(font_name_bold, font_size_base)
                c.drawString(x_inicial, y_positions[5], hecho_en_bold)
                
                # Calcular tamaño adecuado para hecho_en
                c.setFont(font_name, font_size_base)
                ancho_hecho_en = c.stringWidth(hecho_en, font_name, font_size_base)
                tamano_hecho = font_size_base
                
                while ancho_hecho_en > espacio_disponible and tamano_hecho > 5.5:
                    tamano_hecho -= 0.5
                    c.setFont(font_name, tamano_hecho)
                    ancho_hecho_en = c.stringWidth(hecho_en, font_name, tamano_hecho)
                
                c.drawString(x_valores, y_positions[5], hecho_en)
                
                # Pasar a la siguiente página
                c.showPage()
            
            # Guardar el PDF
            c.save()
            return True
            
        except Exception as e:
            import traceback
            traceback.print_exc()  # Imprimir traceback completo para debuggear
            messagebox.showerror("Error", f"Error al generar PDF: {str(e)}")
            return False
    
    def crear_etiquetas_excel(self, archivo_salida, datos):
        """
        Crea un archivo Excel con etiquetas de importación de 76x51mm
        Formato exacto como fue especificado, todo en mayúsculas
        
        Args:
            archivo_salida: Nombre del archivo Excel a generar
            datos: Lista de diccionarios con la información de cada etiqueta
        """
        # Crear un nuevo libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Etiquetas"
        
        # Definir estilos
        borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        no_wrap = Alignment(wrap_text=False, vertical='center', horizontal='left')
        centrado = Alignment(wrap_text=False, vertical='center', horizontal='center')
        
        # Definir diferentes tamaños de fuente
        font_14 = Font(bold=True, size=14)
        font_12 = Font(bold=True, size=12)
        font_11 = Font(bold=False, size=11)
        font_10 = Font(bold=False, size=10)
        font_9 = Font(bold=False, size=9)
        font_8 = Font(bold=False, size=8)
        font_7 = Font(bold=False, size=7)
        
        # Títulos en negrita
        titulo_14 = Font(bold=True, size=14)
        titulo_12 = Font(bold=True, size=12)
        titulo_11 = Font(bold=True, size=11)
        titulo_10 = Font(bold=True, size=10)
        titulo_9 = Font(bold=True, size=9)
        titulo_8 = Font(bold=True, size=8)
        
        # Configurar el ancho de las columnas
        # Columna A para etiquetas, columna B para valores
        ws.column_dimensions['A'].width = 17  # Para títulos/etiquetas
        ws.column_dimensions['B'].width = 40  # Para valores
        ws.column_dimensions['C'].width = 2   # Separador
        ws.column_dimensions['D'].width = 17  # Para títulos/etiquetas
        ws.column_dimensions['E'].width = 40  # Para valores
        
        # Configurar márgenes de página
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        
        # Crear etiquetas
        fila_actual = 1
        col_actual = 1
        
        for etiqueta in datos:
            # Determinar valores para esta etiqueta
            cantidad_contenido = etiqueta.get('cantidad_contenido', 1)
            if cantidad_contenido == 1:
                contenido = f"{cantidad_contenido} PIEZA"
            else:
                contenido = f"{cantidad_contenido} PIEZAS"
            
            descripcion = etiqueta['descripcion'].upper()
            hecho_en = etiqueta['hecho_en'].upper()
            
            # Función para seleccionar fuente según longitud del texto
            def seleccionar_fuente(texto, es_titulo=False):
                longitud = len(texto)
                if es_titulo:
                    if longitud <= 20:
                        return titulo_12
                    elif longitud <= 30:
                        return titulo_10
                    else:
                        return titulo_9
                else:
                    if longitud <= 20:
                        return font_11
                    elif longitud <= 30:
                        return font_10
                    elif longitud <= 40:
                        return font_9
                    elif longitud <= 50:
                        return font_8
                    else:
                        return font_7
            
            # Determinar celda de inicio
            inicio_fila = fila_actual
            inicio_col = col_actual
            
            # 1. IMPORTADOR
            celda_importador = ws.cell(row=inicio_fila, column=inicio_col)
            valor_importador = ws.cell(row=inicio_fila, column=inicio_col + 1)
            celda_importador.value = "IMPORTADOR:"
            celda_importador.font = titulo_12
            celda_importador.alignment = no_wrap
            valor_importador.value = "MOTORMAN DE BAJA CALIFORNIA SA DE CV"
            valor_importador.font = seleccionar_fuente("MOTORMAN DE BAJA CALIFORNIA SA DE CV")
            valor_importador.alignment = no_wrap
            
            # 2. Dirección línea 1
            celda_dir1 = ws.cell(row=inicio_fila + 1, column=inicio_col)
            ws.merge_cells(start_row=inicio_fila + 1, start_column=inicio_col, 
                          end_row=inicio_fila + 1, end_column=inicio_col + 1)
            celda_dir1.value = "MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,"
            celda_dir1.font = seleccionar_fuente("MARISCAL SUCRE 6738 LA CIENEGA PONIENTE,")
            celda_dir1.alignment = no_wrap
            
            # 3. Dirección línea 2
            celda_dir2 = ws.cell(row=inicio_fila + 2, column=inicio_col)
            ws.merge_cells(start_row=inicio_fila + 2, start_column=inicio_col, 
                          end_row=inicio_fila + 2, end_column=inicio_col + 1)
            celda_dir2.value = "TIJUANA, B.C. 22114 RFC: MBC210723RP9"
            celda_dir2.font = seleccionar_fuente("TIJUANA, B.C. 22114 RFC: MBC210723RP9")
            celda_dir2.alignment = no_wrap
            
            # 4. DESCRIPCION
            celda_desc = ws.cell(row=inicio_fila + 3, column=inicio_col)
            valor_desc = ws.cell(row=inicio_fila + 3, column=inicio_col + 1)
            celda_desc.value = "DESCRIPCION:"
            celda_desc.font = titulo_12
            celda_desc.alignment = no_wrap
            valor_desc.value = descripcion
            valor_desc.font = seleccionar_fuente(descripcion)
            valor_desc.alignment = no_wrap
            
            # 5. CONTENIDO
            celda_cont = ws.cell(row=inicio_fila + 4, column=inicio_col)
            valor_cont = ws.cell(row=inicio_fila + 4, column=inicio_col + 1)
            celda_cont.value = "CONTENIDO:"
            celda_cont.font = titulo_12
            celda_cont.alignment = no_wrap
            valor_cont.value = contenido
            valor_cont.font = seleccionar_fuente(contenido)
            valor_cont.alignment = no_wrap
            
            # 6. HECHO EN
            celda_hecho = ws.cell(row=inicio_fila + 5, column=inicio_col)
            valor_hecho = ws.cell(row=inicio_fila + 5, column=inicio_col + 1)
            celda_hecho.value = "HECHO EN:"
            celda_hecho.font = titulo_12
            celda_hecho.alignment = no_wrap
            valor_hecho.value = hecho_en
            valor_hecho.font = seleccionar_fuente(hecho_en)
            valor_hecho.alignment = no_wrap
            
            # Aplicar bordes a todas las celdas
            for r in range(inicio_fila, inicio_fila + 6):
                for c in range(inicio_col, inicio_col + 2):
                    celda = ws.cell(row=r, column=c)
                    # Bordes para primera y última fila
                    if r == inicio_fila:
                        celda.border = Border(
                            left=Side(style='thin') if c == inicio_col else Side(style=None),
                            right=Side(style='thin') if c == inicio_col + 1 else Side(style=None),
                            top=Side(style='thin'),
                            bottom=Side(style=None)
                        )
                    elif r == inicio_fila + 5:
                        celda.border = Border(
                            left=Side(style='thin') if c == inicio_col else Side(style=None),
                            right=Side(style='thin') if c == inicio_col + 1 else Side(style=None),
                            top=Side(style=None),
                            bottom=Side(style='thin')
                        )
                    else:
                        celda.border = Border(
                            left=Side(style='thin') if c == inicio_col else Side(style=None),
                            right=Side(style='thin') if c == inicio_col + 1 else Side(style=None),
                            top=Side(style=None),
                            bottom=Side(style=None)
                        )
            
            # Ajustar altura de filas para etiqueta 51mm (dividido en 6 filas)
            for r in range(inicio_fila, inicio_fila + 6):
                ws.row_dimensions[r].height = 8.5  # Aproximadamente 51mm/6
            
            # Avanzar a la siguiente posición
            if col_actual >= 4:  # Máximo 2 etiquetas por fila (cada etiqueta usa 2 columnas + 1 separador)
                col_actual = 1
                fila_actual += 6  # Saltar a la siguiente fila de etiquetas
            else:
                col_actual += 3  # Avanzar a la siguiente etiqueta en la misma fila
        
        # Guardar el archivo
        wb.save(archivo_salida)

if __name__ == "__main__":
    root = tk.Tk()
    app = GeneradorEtiquetasApp(root)
    root.mainloop()